from tqdm import tqdm
from openpyxl import load_workbook
import csv
from skylark_pairdiff import get_functional_group_change

def main(data_save_dir):
    wb = load_workbook(data_save_dir + '\\'+ r'msdial_result.xlsx')
    ws = wb.active
    
    known_dict = {}
    dot_id_list = []
    for i in tqdm(range(2, ws.max_row+1)):
        smiles = ws.cell(i, 5).value
        if smiles != "unknown":
            dot_id = int(ws.cell(i, 1).value)
            known_dict[dot_id] = smiles
            dot_id_list.append(dot_id)
    
    final_list = []   
    # 计算真实结构差
    for id_a in tqdm(dot_id_list):
        for id_b in dot_id_list:
            try:
                if id_a != id_b:
                    smiles_a = known_dict[id_a]
                    smiles_b = known_dict[id_b]
                    replace_list = get_functional_group_change(smiles_a, smiles_b)
                    if replace_list != "None":
                        replaced, replace = replace_list[0], replace_list[1]
                        if replaced != 'unknown':
                            replace_str = replaced + "-" + replace
                            if id_a > id_b:
                                id_a, id_b = id_b, id_a
                            final_list.append([id_a, id_b, replace_str])
            except:
                pass
    
    # 写入csv
    with open(data_save_dir + '\\'+ r'真实结构差.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(final_list)


if __name__ == "__main__":
    main()