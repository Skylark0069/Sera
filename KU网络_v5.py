import pubchempy as pcp
import csv
import os
import pandas as pd
import re
import rdkit
from rdkit import Chem
from openpyxl import load_workbook
from tqdm import tqdm
from AutoML_model_use_function_v5 import *
import multiprocessing
from multiprocessing import Process
import joblib

# 标准化SMILES，失败返回'None'
def SMILES_to_standard(SMILES):
    try:
        mol = Chem.MolFromSmiles(SMILES)
        standard_SMILES = Chem.MolToSmiles(mol)
        return standard_SMILES
    except:
        return 'None'

def split_mission(known_list, unknown_list, nist_list, data_save_dir, data_save_name, cpu_num, sim_threshold):
    total_num = len(unknown_list)
    mission_num = int(cpu_num)
    split_range_num = total_num // mission_num
    # 将dot_id_list划分为mission_num份, 输出划分区间的索引
    split_mission_list = []
    for i in range(mission_num):
        if i == (mission_num-1):
            start_index, end_index = i*split_range_num, total_num
        else:
            start_index, end_index = i*split_range_num, (i+1)*split_range_num
        split_mission_list.append((start_index, end_index, known_list, unknown_list, nist_list, data_save_dir, data_save_name, sim_threshold))
    return split_mission_list

mass_label_dict = joblib.load(r'dict\\模型标签字典.joblib')
mass_range_dict = joblib.load(r'dict\\模型质量范围字典.joblib')

def get_ku_net(start_end):
    final_list = []
    start_index, end_index, known_list, unknown_list, nist_list, data_save_dir, data_save_name, sim_threshold = start_end
    for a in tqdm(range(start_index, end_index)):
        unknown_id, unknown_smiles, unknown_mass, unknown_msms = unknown_list[a][0], 'unknown', float(unknown_list[a][2]), unknown_list[a][3]
        for b in range(len(known_list)):
            try:
                known_id, known_smiles, known_mass, known_msms = known_list[b][0], known_list[b][1], float(known_list[b][2]), known_list[b][3]
                if abs(unknown_mass - known_mass) < 100:
                    sim = MS_sim_type2(unknown_msms, known_msms)
                    if sim > sim_threshold:
                        model_id = get_model_id(unknown_mass, known_mass)
                        
                        if model_id == 'unknown':
                            continue
                        else:
                            massdiff = abs(unknown_mass - known_mass)
                            mass0, mass1 = mass_range_dict[model_id][0], mass_range_dict[model_id][1]
                            
                            if mass0 <= massdiff <= mass1:
                                mz0, int0 = get_mz_inten_list_type2(unknown_msms)
                                mz1, int1 = get_mz_inten_list_type2(known_msms)

                                if known_mass < unknown_mass:
                                    mz0, int0, mz1, int1 = mz1, int1, mz0, int0
                                    pairdiff1, pairdiff2 = get_pairdiff(model_id, mz0, int0, mz1, int1)
                                    if pairdiff1 != 'unknown':
                                        TF = is_substructure(pairdiff1, known_smiles)
                                        if TF == True:
                                            pairdiff1, pairdiff2 = pairdiff1, pairdiff2
                                        else:
                                            label_list = mass_label_dict[model_id][:-1]
                                            for label_pair in label_list:
                                                if is_substructure(label_pair[0], known_smiles) == "True":
                                                    pairdiff1, pairdiff2 = label_pair[0], label_pair[1]
                                                    break
                                                else:
                                                    pairdiff1, pairdiff2 = "unknown", "unknown"
                                    else:
                                        pairdiff1, pairdiff2 = "unknown", "unknown"

                                else:
                                    pairdiff1, pairdiff2 = get_pairdiff(model_id, mz0, int0, mz1, int1)
                                    if pairdiff1 != 'unknown':
                                        TF = is_substructure(pairdiff2, known_smiles)
                                        if TF == True:
                                            pairdiff1, pairdiff2 = pairdiff1, pairdiff2
                                        else:
                                            label_list = mass_label_dict[model_id][:-1]
                                            for label_pair in label_list:
                                                if is_substructure(label_pair[1], known_smiles) == "True":
                                                    pairdiff1, pairdiff2 = label_pair[0], label_pair[1]
                                                    break
                                                else:
                                                    pairdiff1, pairdiff2 = "unknown", "unknown"
                                
                                if pairdiff1 != 'unknown':
                                    pairdiff = pairdiff1 + '-' + pairdiff2
                                    # print(pairdiff)
                                    if known_mass < unknown_mass:
                                        final_list.append([known_id, known_smiles, unknown_id, unknown_smiles, sim, pairdiff])
                                    else:
                                        final_list.append([unknown_id, unknown_smiles, known_id, known_smiles, sim, pairdiff])
                                    print(final_list[-1])
                    else:
                        continue
            except:
                print('error1')
                continue

        for c in range(len(nist_list)):
            try:
                nist_id, nist_smiles, nist_mass, nist_msms = 'nist', nist_list[c][0], float(nist_list[c][1]), nist_list[c][2]
                if abs(unknown_mass - nist_mass) < 100:
                    sim = MS_sim_type2(unknown_msms, nist_msms)
                    if sim > sim_threshold:
                        model_id = get_model_id(unknown_mass, nist_mass)

                        if model_id == 'unknown':
                            continue
                        else:
                            massdiff = abs(unknown_mass - nist_mass)
                            mass0, mass1 = mass_range_dict[model_id][0], mass_range_dict[model_id][1]

                            if mass0 <= massdiff <= mass1:
                                mz0, int0 = get_mz_inten_list_type2(unknown_msms)
                                mz1, int1 = get_mz_inten_list_type2(nist_msms)
                                if nist_mass < unknown_mass:
                                    mz0, int0, mz1, int1 = mz1, int1, mz0, int0
                                    pairdiff1, pairdiff2 = get_pairdiff(model_id, mz0, int0, mz1, int1)
                                    if pairdiff1 != 'unknown':
                                        TF = is_substructure(pairdiff1, nist_smiles)
                                        if TF == True:
                                            pairdiff1, pairdiff2 = pairdiff1, pairdiff2
                                        else:
                                            label_list = mass_label_dict[model_id][:-1]
                                            for label_pair in label_list:
                                                if is_substructure(label_pair[0], nist_smiles) == "True":
                                                    pairdiff1, pairdiff2 = label_pair[0], label_pair[1]
                                                    break
                                                else:
                                                    pairdiff1, pairdiff2 = "unknown", "unknown"
                                    else:
                                        pairdiff1, pairdiff2 = "unknown", "unknown"

                                else:
                                    pairdiff1, pairdiff2 = get_pairdiff(model_id, mz0, int0, mz1, int1)
                                    if pairdiff1 != 'unknown':
                                        TF = is_substructure(pairdiff2, nist_smiles)
                                        if TF == True:
                                            pairdiff1, pairdiff2 = pairdiff1, pairdiff2
                                        else:
                                            label_list = mass_label_dict[model_id][:-1]
                                            for label_pair in label_list:
                                                if is_substructure(label_pair[1], nist_smiles) == "True":
                                                    pairdiff1, pairdiff2 = label_pair[0], label_pair[1]
                                                    break
                                                else:
                                                    pairdiff1, pairdiff2 = "unknown", "unknown"
                            
                            
                                if pairdiff1 != 'unknown':
                                    pairdiff = pairdiff1 + '-' + pairdiff2 
                                    # print(pairdiff)
                                    if nist_mass < unknown_mass:
                                        final_list.append([nist_id, nist_smiles, unknown_id, unknown_smiles, sim, pairdiff])
                                    else:
                                        final_list.append([unknown_id, unknown_smiles, nist_id, nist_smiles, sim, pairdiff])
                                    print(final_list[-1])
                    else:           
                         continue
            except:
                print('error2')
                continue
            
    # 将final_list写入csv
    with open(data_save_dir + '\\'+ 'KU_net.csv', 'a', newline='') as f:
        writer = csv.writer(f)
        for i in range(len(final_list)):
            writer.writerow(final_list[i])


def main(data_save_dir, data_save_name, cpu_num, sim_threshold):
    cpu_num = int(cpu_num)
    wb = load_workbook(data_save_dir + '\\'+ r'msdial_result.xlsx')
    ws = wb.active
    known_list = []
    unknown_list = []
    for i in tqdm(range(2, ws.max_row+1)):
        dot_id, smiles, mass, msms = ws.cell(i,1).value, ws.cell(i,5).value, ws.cell(i,4).value, ws.cell(i,7).value
        if msms != None:
            if smiles == 'unknown':
                unknown_list.append([dot_id, smiles, mass, msms])
            else:
                smiles = SMILES_to_standard(smiles)
                if smiles != 'None':
                    known_list.append([dot_id, smiles, mass, msms])
                else:
                    unknown_list.append([dot_id, 'unknown', mass, msms])     
    
    wb0 = load_workbook(r'NIST_POS_1.xlsx')
    ws0 = wb0.active
    nist_list = []
    for i in tqdm(range(1, ws0.max_row+1)):
        try:
            smiles, mass, msms_str = ws0.cell(i,1).value, ws0.cell(i,2).value, ws0.cell(i,3).value
            smiles = SMILES_to_standard(smiles)
            msms_str = msms_str.replace('[', '').replace(']', '').replace(' ', '')
            msms_str_list = msms_str.split(',')
            # 索引为奇数的是mz，索引为偶数的是int
            mz_list, int_list = [], []
            for j in range(len(msms_str_list)):
                if j % 2 == 0:
                    mz_list.append(float(msms_str_list[j]))
                else:
                    int_list.append(float(msms_str_list[j]))
            final_msms_str = ''
            for k in range(len(mz_list)):
                final_msms_str += str(mz_list[k]) + ' ' + str(int_list[k]) + ';'
            final_msms_str = final_msms_str[:-1]
            nist_list.append([smiles, mass, final_msms_str])
        except:
            continue
    
    split_mission_list = split_mission(known_list, unknown_list, nist_list, data_save_dir, data_save_name, cpu_num, sim_threshold)
    pool = multiprocessing.Pool(cpu_num)
    rel = pool.map(get_ku_net, split_mission_list)
    pool.close()
    pool.join()
    

if __name__ == '__main__':
    main()