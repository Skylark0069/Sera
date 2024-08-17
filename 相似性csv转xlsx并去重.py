# 将csv文件转成xlsx文件
import csv
import os
import sys
import pandas as pd
from openpyxl import Workbook, load_workbook
from tqdm import tqdm

def csv2xlsx(csv_file, xlsx_file):
    wb = Workbook()
    ws = wb.active
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)
    # 表格的第一列和第二列数据类型转化为int, 第三列转化为float
    for row in ws.iter_rows(min_row=1, max_col=3):
        row[0].value = int(row[0].value)
        row[1].value = int(row[1].value)
        row[2].value = float(row[2].value)
    # 加入表头
    ws.insert_rows(0)
    ws.cell(1,1).value, ws.cell(1,2).value, ws.cell(1,3).value = 'dot_id0', 'dot_id1', 'sim'
    wb.save(xlsx_file)
    # 对xlsx进行去重, 去除两行数据相同的行中的一行
    df1 = pd.read_excel(xlsx_file)
    df2 = df1.drop_duplicates(subset=['dot_id0', 'dot_id1'], keep='first')
    df2.to_excel(xlsx_file, index=False)
    # 删除表头
    wb = load_workbook(xlsx_file)
    ws = wb.active
    ws.delete_rows(1)
    wb.save(xlsx_file)
    # 删除csv
    os.remove(csv_file)

def main(data_save_dir):
    csv2xlsx(data_save_dir + '\\'+ r'相似性.csv', data_save_dir + '\\'+ r'相似性.xlsx')

if __name__ == '__main__':
    main()