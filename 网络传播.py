from openpyxl import load_workbook
import csv
from tqdm import tqdm
import multiprocessing
from multiprocessing import Process
from AutoML_model_use_function_v5 import *


# 获取母分子id子分子id-相似性字典, 母分子id子分子id-结构差字典, 母分子id-[子分子id]字典
def get_3dict(data_save_dir):
    wb = load_workbook(data_save_dir)
    ws = wb.active
    keys_list = []
    dot_ids_pairdiff_dict = {}
    dot_ids_sim_dict = {}
    for i in tqdm(range(1, ws.max_row+1)):
        dot_id1, dot_id2, sim, diff = ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,3).value, ws.cell(i,4).value
        keys_list.append(dot_id1)
        keys_list.append(dot_id2)
        dot_ids_sim_dict[str(dot_id1)+'-'+str(dot_id2)] = sim
        dot_ids_pairdiff_dict[str(dot_id1)+'-'+str(dot_id2)] = diff
    keys_list = list(set(keys_list))
    pairdiff_id_dict = {}
    for dot_id0 in tqdm(keys_list):
        dot_id0_list = []
        for i in range(1, ws.max_row+1):
            dot_id1, dot_id2 = ws.cell(i,1).value, ws.cell(i,2).value
            if dot_id0 == dot_id1:
                dot_id0_list.append(dot_id2)
            elif dot_id0 == dot_id2:
                dot_id0_list.append(dot_id1)
        dot_id0_list = list(set(dot_id0_list))
        pairdiff_id_dict[dot_id0] = dot_id0_list
    return dot_ids_sim_dict, dot_ids_pairdiff_dict, pairdiff_id_dict

# 每次迭代
def get_pairdiff_side(dot_id_mass_dict, known_dot_list, unknown_dot_list, dot_ids_sim_dict, 
                      dot_ids_pairdiff_dict, pairdiff_id_dict, dot_id_smiles_dict):
    known_dot_list1, unknown_dot_list1 = known_dot_list.copy(), unknown_dot_list.copy()
    pairdiff_side_list = []
    for known_dot_id in tqdm(known_dot_list):
        if known_dot_id in pairdiff_id_dict.keys():
            pairdiff_dot_list = pairdiff_id_dict[known_dot_id]
            for pairdiff_dot_id in pairdiff_dot_list:
                try:
                    A_smiles = dot_id_smiles_dict[known_dot_id]
                except:
                    A_smiles = 'unknown'
                try:
                    B_smiles = dot_id_smiles_dict[pairdiff_dot_id]
                except:
                    B_smiles = 'unknown'
                if pairdiff_dot_id in unknown_dot_list:
                    known_dot_list1.append(pairdiff_dot_id)
                    try:
                        unknown_dot_list1.remove(pairdiff_dot_id)
                    except:
                        continue
                    try:
                        known_dot_id_mass, pairdiff_dot_id_mass = dot_id_mass_dict[known_dot_id], dot_id_mass_dict[pairdiff_dot_id]
                        if known_dot_id_mass < pairdiff_dot_id_mass:
                            sim = dot_ids_sim_dict[str(known_dot_id)+'-'+str(pairdiff_dot_id)]
                            pairdiff = dot_ids_pairdiff_dict[str(known_dot_id)+'-'+str(pairdiff_dot_id)]
                            pairdiff_side_list.append([known_dot_id, A_smiles, pairdiff_dot_id, B_smiles, sim, pairdiff])
                        else:
                            sim = dot_ids_sim_dict[str(pairdiff_dot_id)+'-'+str(known_dot_id)]
                            pairdiff = dot_ids_pairdiff_dict[str(pairdiff_dot_id)+'-'+str(known_dot_id)]
                            pairdiff_side_list.append([pairdiff_dot_id, B_smiles, known_dot_id, A_smiles, sim, pairdiff])
                    except:
                        continue
    known_dot_list1 = list(set(known_dot_list1))
    unknown_dot_list1 = list(set(unknown_dot_list1))
    return pairdiff_side_list, known_dot_list1, unknown_dot_list1

# 网络传播
def net_spread(dot_id_mass_dict, dot_id_smiles_dict, data_save_dir, save_data_dir, count):
    # data_dir_A: msidal_result.xlsx; data_dir_B: 结构差.xlsx
    data_dir_A = data_save_dir + r'\\msdial_result.xlsx'
    data_dir_B = data_save_dir + r'\\结构差.xlsx'
    final_list = []
    known_dot_list, unknown_dot_list = get_known_unknown_list(data_dir_A)
    #print('获取3个字典')
    dot_ids_sim_dict, dot_ids_pairdiff_dict, pairdiff_id_dict = get_3dict(data_dir_B)
    a = 0
    while a < count:
        start_len = len(final_list)
        #print('第' + str(a+1) + '次传播...') 
        pairdiff_list, known_dot_list, unknown_dot_list = get_pairdiff_side(dot_id_mass_dict, known_dot_list, unknown_dot_list, 
                                                                            dot_ids_sim_dict, dot_ids_pairdiff_dict, 
                                                                            pairdiff_id_dict, dot_id_smiles_dict)
        pairdiff_list = add_element_to_list(pairdiff_list, a)
        final_list.extend(pairdiff_list)
        a = a + 1
        #print('已知节点数：' + str(len(known_dot_list)))
        #print('未知节点数：' + str(len(unknown_dot_list)))
        if len(final_list) == start_len:
            break   # 如果迭代后没有出现新的边，则跳出循环
    f = open(save_data_dir,'w',encoding='utf-8',newline='')
    csv_writer = csv.writer(f)
    # 写入表头
    csv_writer.writerow(['DotID1','SMILES1','DotID2','SMILES2','sim','pairdiff','spread_round'])
    for i in final_list:
        csv_writer.writerow(i) 

def main(data_save_dir, data_save_name, spread_round):
    spread_round = int(spread_round)
    data_dir = data_save_dir + '\\'+ r'msdial_result.xlsx'
    wb1 = load_workbook(data_dir)
    ws1 = wb1.active
    # 获取dot_id和mass的字典
    # 获取dot_id和smiles的字典
    dot_id_mass_dict = {}
    dot_id_smiles_dict = {}
    for i in range(2, ws1.max_row+1):
        dot_id, mass, smiles = ws1.cell(i,1).value, ws1.cell(i,4).value, ws1.cell(i,5).value
        dot_id_mass_dict[dot_id] = mass
        if smiles != 'unknown':
            dot_id_smiles_dict[dot_id] = smiles
    dot_ids_sim_dict, dot_ids_pairdiff_dict, pairdiff_id_dict = get_3dict(data_save_dir+'\\结构差.xlsx')
    net_spread(dot_id_mass_dict, dot_id_smiles_dict, data_save_dir, 
               data_save_dir + '\\'+ r'网络传播.csv', spread_round)

if __name__ == '__main__':
    main()