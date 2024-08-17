# 将csv文件转成xlsx文件
import csv
import os
import sys
from openpyxl import Workbook

def csv2xlsx(csv_file, xlsx_file):
    wb = Workbook()
    ws = wb.active
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    for row in ws.iter_rows(min_row=1, max_col=6):
        if "\ufeff" in row[0].value:
            row[0].value = row[0].value.replace("\ufeff", "")
        row[0].value = int(row[0].value)
        row[1].value = str(row[1].value)
        row[2].value = int(row[2].value)
        row[3].value = int(row[3].value)
        row[4].value = int(row[4].value)
        row[5].value = str(row[5].value)
    wb.save(xlsx_file)
    # 删除csv文件
    os.remove(csv_file)

def csv2xlsx2(csv_file, xlsx_file):
    wb = Workbook()
    ws = wb.active
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    for row in ws.iter_rows(min_row=1, max_col=5):
        try:
            row[0].value = int(row[0].value)
        except:
            row[0].value = str(row[0].value)
        try:
            row[2].value = int(row[2].value)
        except:
            row[2].value = str(row[2].value)
        row[1].value = str(row[1].value)
        row[3].value = str(row[3].value)
        row[4].value = str(row[4].value)
    wb.save(xlsx_file)
    # 删除csv文件
    os.remove(csv_file)
        
def main(data_save_dir, data_save_name):
    csv2xlsx(data_save_dir + '\\'+ data_save_name + 'Structural Results.csv', data_save_dir + '\\'+ data_save_name + ' Structural Results.xlsx')
    csv2xlsx2(data_save_dir + '\\'+ '参与成网边.csv', data_save_dir + '\\'+ '参与成网边.xlsx')

if __name__ == '__main__':
    main()