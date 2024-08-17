import pubchempy as pcp
from openpyxl import load_workbook
import csv
from tqdm import tqdm
import re
import os
import rdkit
from rdkit import Chem
from rdkit.Chem import rdFMCS
from rdkit.Chem import AllChem
from rdkit.Chem import Draw
from rdkit.Chem import rdMolHash
from rdkit.Chem import Descriptors
from skylark_pairdiff import get_functional_group_change
import multiprocessing
from multiprocessing import Process

# 标准化SMILES，失败返回'None'
def SMILES_to_standard(SMILES):
    try:
        mol = Chem.MolFromSmiles(SMILES)
        standard_SMILES = Chem.MolToSmiles(mol)
        return standard_SMILES
    except:
        return 'None'

# 将SMILES转为化学式字典，失败返回'None'
def SMILES_to_formula_dict(SMILES):
    try:
        # 把SMILES转为显式H版本
        mol = Chem.MolFromSmiles(SMILES)
        mol = Chem.AddHs(mol)
        SMILES = Chem.MolToSmiles(mol)
        # 元素列表
        element_list = ['H', 'C', 'c', 'N', 'O', 
                        'F', 'Na', 'Mg', 'Al', 'Si', 'P', 'S', 'Cl', 
                        'K', 'Ca', 'Cr', 'Mn', 'Fe', 'Cu', 'Zn',
                        'Br', 'Ag', 'Cd', 'I','Ba']
        # 获取出现的元素的{元素：数量}字典
        element_dict = {}
        for element in element_list:
            element_num = SMILES.count(element)
            if element_num != 0:
                element_dict[element] = element_num
        # 修正元素数量
        for element in element_dict:
            if element == 'C':
                element_dict[element] += SMILES.count('c')
                element_dict[element] -= SMILES.count('Cl')+SMILES.count('Ca')+SMILES.count('Cr')+SMILES.count('Cu')+SMILES.count('Cd')
            elif element == 'F':
                element_dict[element] -= SMILES.count('Fe')
            elif element == 'S':
                element_dict[element] -= SMILES.count('Si')
            elif element == 'N':
                element_dict[element] -= SMILES.count('Na')
        # 如果有c的键，则从字典中删除c的键值对
        if 'c' in element_dict:
            del element_dict['c']
        # 如果没有C，则添加C的键值对
        if 'C' not in element_dict:
            element_dict['C'] = SMILES.count('c')
        # print(element_dict)
        return element_dict
    except:
        return 'None'

# 判断SMILES是否正确，正确返回1，否则返回0
def SMILES_tf(SMILES):
    try:
        mol = Chem.MolFromSmiles(SMILES)
        Draw.MolToFile(mol, r'trash\\' + SMILES + '.png', size=(10, 10))
        os.remove( r'trash\\' + SMILES + '.png')
        return 1
    except:
        return 0

# 根据分子式搜索SMILES，返回候选SMILES列表，失败返回'None'
def search_pubchem_SMILES(formula):
    try:
        molecular_list = pcp.get_compounds(formula, 'formula')
        SMILES_list = []
        for molecular in molecular_list:
            standard_SMILES = SMILES_to_standard(molecular.canonical_smiles)
            if standard_SMILES != 'None' and '.' not in standard_SMILES and '[' not in standard_SMILES:
                SMILES_list.append(standard_SMILES)
        if len(SMILES_list) == 0:
            return 'None'
        else:
            SMILES_list = list(set(SMILES_list))
            return SMILES_list
    except:
        return 'None'

# 根据SMILES计算分子质量
def get_molecular_mass(SMILES):
    mol = Chem.MolFromSmiles(SMILES)
    return Descriptors.ExactMolWt(mol)

# 使用rdkit, 判断SMILES代表的分子是否含有环状结构
def SMILES_has_ring(SMILES):
    mol = Chem.MolFromSmiles(SMILES)
    if mol.HasSubstructMatch(Chem.MolFromSmarts('a')):
        return 1
    else:
        return 0

# 分子式加减计算，输入分子式字典，返回分子式，失败返回'None'
def formula_add(smiles, add_smiles1, add_smiles2, add_type):
    smiles_dict = SMILES_to_formula_dict(smiles)
    add_smiles_dict1 = SMILES_to_formula_dict(add_smiles1)
    add_smiles_dict2 = SMILES_to_formula_dict(add_smiles2)
    if smiles_dict == 'None' or add_smiles_dict1 == 'None' or add_smiles_dict2 == 'None':
        return 'None'
    else:
        # add_type = 0: smiles为母体：smiles - add_smiles1 + add_smiles2 = final_smiles
        # add_type = 1: smiles为子体：smiles + add_smiles1 - add_smiles2 = final_smiles
        if add_type == 0:
            for element in add_smiles_dict1:
                if element in smiles_dict:
                    smiles_dict[element] -= add_smiles_dict1[element]
                else:
                    return 'None'
            for element in add_smiles_dict2:
                if element in smiles_dict:
                    smiles_dict[element] += add_smiles_dict2[element]
                else:
                    smiles_dict[element] = add_smiles_dict2[element]
        elif add_type == 1:
            for element in add_smiles_dict1:
                if element in smiles_dict:
                    smiles_dict[element] += add_smiles_dict1[element]
                else:
                    smiles_dict[element] = add_smiles_dict1[element]
            for element in add_smiles_dict2:
                if element in smiles_dict:
                    smiles_dict[element] -= add_smiles_dict2[element]
                else:
                    return 'None'
        else:
            return 'None'
        final_formula = ''
        for element in smiles_dict:
            # 让分子式中的元素按照C、H在前，其他元素按照字母顺序排列
            if smiles_dict[element] != 0:
                if element == 'C':
                    final_formula = 'C' + str(smiles_dict[element]) + final_formula
                elif element == 'H':
                    final_formula = 'H' + str(smiles_dict[element]) + final_formula
                else:
                    if smiles_dict[element] == 1:
                        final_formula += element
                    else:
                        final_formula += element + str(smiles_dict[element])
        return final_formula

# 让分子式中的H加2或者减2，返回分子式
def formula_H_add2(formula, type):
    try:
        # 找到H的位置
        H_index = formula.find('H')
        # 获取H之后下一个字母的位置
        for i in range(H_index + 1, len(formula)):
            if formula[i].isalpha():
                H_next_index = i
                break
        # 获取H_index和H_next_index之间的数字
        H_num = int(formula[H_index + 1:H_next_index])
        # type = +: H加2
        # type = -: H减2
        if type == '+':
            H_num += 2
        elif type == '-':
            H_num -= 2
        # 重新拼接分子式
        if H_num == 1:
            formula = formula[:H_index] + formula[H_next_index:]
        else:
            formula = formula[:H_index + 1] + str(H_num) + formula[H_next_index:]
        return formula
    except:
        return 'None'

# 根据结构差和母体或子体，返回候选结构差列表，失败返回'None'
def search_pairdiff_SMILES(pairdiff, known_smiles, search_type):
    if pairdiff != "Isomers":
        pairdiff_a, pairdiff_b = pairdiff.split('-')
        # search_type = 0: known_smiles为母体，搜索子体
        # search_type = 1: known_smiles为子体，搜索母体
        if search_type == 0:
            final_formula = formula_add(known_smiles, pairdiff_a, pairdiff_b, 0)
            final_formula_ring0 = formula_H_add2(final_formula, '+')
            final_formula_ring1 = formula_H_add2(final_formula, '-')
        elif search_type == 1:
            final_formula = formula_add(known_smiles, pairdiff_a, pairdiff_b, 1)
            final_formula_ring0 = formula_H_add2(final_formula, '+')
            final_formula_ring1 = formula_H_add2(final_formula, '-')
        if final_formula == 'None':
            return 'None'
        else:
            final_SMILES_list = []
            SMILES_list = search_pubchem_SMILES(final_formula)
            SMILES_list_ring0 = search_pubchem_SMILES(final_formula_ring0)
            SMILES_list_ring1 = search_pubchem_SMILES(final_formula_ring1)
            if SMILES_list != 'None':
                if SMILES_list_ring0 != 'None':
                    SMILES_list.extend(SMILES_list_ring0)
                if SMILES_list_ring1 != 'None':
                    SMILES_list.extend(SMILES_list_ring1)
            else:
                if SMILES_list_ring0 != 'None':
                    SMILES_list = SMILES_list_ring0
                    if SMILES_list_ring1 != 'None':
                        SMILES_list.extend(SMILES_list_ring1)
                else:
                    if SMILES_list_ring1 != 'None':
                        SMILES_list = SMILES_list_ring1
                    else:
                        SMILES_list = 'None'
            # print(SMILES_list)
            print(len(SMILES_list))
            if SMILES_list != 'None':
                if search_type == 0:
                    for unknown_smiles in SMILES_list:
                        unknown_smiles = SMILES_to_standard(unknown_smiles)
                        try:
                            pairdiff_aa, pairdiff_bb = get_functional_group_change(known_smiles, unknown_smiles)
                            pairdiff_aa, pairdiff_bb = SMILES_to_standard(pairdiff_aa), SMILES_to_standard(pairdiff_bb)
                            if (pairdiff_aa == pairdiff_a and pairdiff_bb == pairdiff_b) or (pairdiff_aa == pairdiff_b and pairdiff_bb == pairdiff_a):
                                final_SMILES_list.append(unknown_smiles)
                        except:
                            continue
                elif search_type == 1:
                    for unknown_smiles in SMILES_list:
                        unknown_smiles = SMILES_to_standard(unknown_smiles)
                        try:
                            pairdiff_aa, pairdiff_bb = get_functional_group_change(unknown_smiles, known_smiles)
                            pairdiff_aa, pairdiff_bb = SMILES_to_standard(pairdiff_aa), SMILES_to_standard(pairdiff_bb)
                            if (pairdiff_aa == pairdiff_a and pairdiff_bb == pairdiff_b) or (pairdiff_aa == pairdiff_b and pairdiff_bb == pairdiff_a):
                                final_SMILES_list.append(unknown_smiles)
                        except:
                            continue
            if len(final_SMILES_list) == 0:
                return 'None'
            else:
                return final_SMILES_list
    else:
        return 'None'


def split_mission(mother_known_dot_pair_list, son_known_dot_pair_list, dot_list, cpu_num, data_save_dir, data_save_name):
    total_num = len(dot_list)
    mission_num = int(cpu_num)
    split_range_num = total_num // mission_num
    split_mission_list = []
    for i in range(mission_num):
        if i == (mission_num-1):
            start_index, end_index = i*split_range_num, total_num
        else:
            start_index, end_index = i*split_range_num, (i+1)*split_range_num
        split_mission_list.append((start_index, end_index, mother_known_dot_pair_list, 
                                   son_known_dot_pair_list, dot_list, cpu_num, data_save_dir, data_save_name))
    return split_mission_list 

def search_smiles(start_end):
    start_index, end_index, mother_known_dot_pair_list, son_known_dot_pair_list, dot_list, cpu_num, data_save_dir, data_save_name = start_end
    final_list = []
    # 参与结构鉴定的边
    side_list = []
    for dot_id0 in tqdm(dot_list[start_index:end_index]):
        # 获取候选结构差列表
        candidate_result_list = []
        for a in mother_known_dot_pair_list:
            dot_id1, smiles1, dot_id2, smiles2, pairdiff = a[0], a[1], a[2], a[3], a[4]
            if dot_id0 == dot_id2:
                SMILES_list = search_pairdiff_SMILES(pairdiff, smiles1, 0)
                # print(SMILES_list)
                if SMILES_list != 'None':
                    SMILES_list = list(set(SMILES_list))
                    candidate_result_list.append(SMILES_list)
                    side_list.append([dot_id1, smiles1, dot_id2, smiles2, pairdiff])
        for b in son_known_dot_pair_list:
            dot_id1, smiles1, dot_id2, smiles2, pairdiff = b[0], b[1], b[2], b[3], b[4]
            if dot_id0 == dot_id1:
                SMILES_list = search_pairdiff_SMILES(pairdiff, smiles2, 1)
                if SMILES_list != 'None':
                    SMILES_list = list(set(SMILES_list))
                    candidate_result_list.append(SMILES_list)
                    side_list.append([dot_id1, smiles1, dot_id2, smiles2, pairdiff])
        side_count = len(candidate_result_list)
        # 展开候选结构差嵌套列表
        candidate_result_list_0 = []
        for i_list in candidate_result_list:
            for i in i_list:
                candidate_result_list_0.append(i)
        # 去重
        candidate_result_list_1 = list(set(candidate_result_list_0))
        # 统计候选结构个数
        for candidate_result in candidate_result_list_1:
            total_count = len(candidate_result_list_1)
            rata = "?"
            if total_count == 1:
                rate = "A"
            elif total_count <= 5:
                rate = "B"
            else:
                rate = "C"
            # 统计candidate_result在candidate_result_list_0中出现的次数
            count = candidate_result_list_0.count(candidate_result)
            final_list.append([dot_id0, candidate_result, count, side_count, total_count, rate])
            print([dot_id0, candidate_result, count, side_count, total_count, rate])
    # final_list写入csv
    with open(data_save_dir + '\\'+ data_save_name + 'Structural Results.csv', 'a', newline='') as f:
        writer = csv.writer(f)
        for i in final_list:
            writer.writerow(i)
    
    with open(data_save_dir + '\\'+ '参与成网边.csv', 'a', newline='') as f:
        writer = csv.writer(f)
        for i in side_list:
            writer.writerow(i)

def main(data_save_dir, data_save_name, cpu_num):
    wb = load_workbook(data_save_dir + '\\'+ r'总表.xlsx')
    ws = wb.active
    # 获取母体已知子体未知的节点对列表和子体已知母体未知的节点对列表
    mother_known_dot_pair_list = []
    son_known_dot_pair_list = []
    dot_list = []
    nist_index = 1
    for i in range(1, ws.max_row+1):
        DotID1, SMILES1, DotID2, SMILES2, pairdiff = ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,3).value, ws.cell(i,4).value, ws.cell(i,6).value
        if SMILES1 != 'unknown' and SMILES2 == 'unknown':
            if DotID1 == 'nist':
                DotID1 = 'nist' + str(nist_index)
                nist_index += 1
            mother_known_dot_pair_list.append([DotID1, SMILES1, DotID2, SMILES2, pairdiff])
        elif SMILES1 == 'unknown' and SMILES2 != 'unknown':
            if DotID2 == 'nist':
                DotID2 = 'nist' + str(nist_index)
                nist_index += 1
            son_known_dot_pair_list.append([DotID1, SMILES1, DotID2, SMILES2, pairdiff])
        if 'nist' not in str(DotID1):
            dot_list.append(DotID1)
        if 'nist' not in str(DotID2):
            dot_list.append(DotID2)
    # 去重
    dot_list = list(set(dot_list))
    # 分配任务
    split_mission_list = split_mission(mother_known_dot_pair_list, son_known_dot_pair_list, dot_list, cpu_num, data_save_dir, data_save_name)
    pool = multiprocessing.Pool(cpu_num)
    rel = pool.map(search_smiles, split_mission_list)
    pool.close()

if __name__ == '__main__':
    main()