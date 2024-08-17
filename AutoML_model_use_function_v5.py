from openpyxl import load_workbook
import numpy as np
import pandas as pd
import csv
import joblib
from tqdm import tqdm
import sklearn
from sklearn.ensemble import ExtraTreesClassifier
from sklearn.ensemble import RandomForestClassifier
from tpot import TPOTClassifier
import rdkit
from rdkit import Chem
from rdkit.Chem import rdMolDescriptors

# 读取模型标签字典
# 0: [['CC', 'CO'], ['C=O', 'CC'], ['C(=O)O', 'CCO'], ['unknown', 'unknown']]
mass_label_dict = joblib.load(r'dict\\模型标签字典.joblib')
model_id_list = list(mass_label_dict.keys())
try:
    new_mass_label_dict_1 = {}
    for key in mass_label_dict.keys():
        a_dict = mass_label_dict[key]
        a_list = []
        for key0 in a_dict.keys():
            pairdiff1, pairdiff2 = a_dict[key0].split('_')
            a_list.append([pairdiff1, pairdiff2])
        new_mass_label_dict_1[key] = a_list
    mass_label_dict = new_mass_label_dict_1
except:
    pass

# 读取模型质量范围字典
# 0: [0, 0.094]
mass_range_dict = joblib.load(r'旧_dict\\模型质量范围字典.joblib')

def get_mz_inten_list(str):
    '''
    获得mz和inten列表
    输入: 50.10775 3776;56.76578 3510;57.0704 14032;69.07013 5666;71.08612 7535;87.50748 4250;103.07556 11173
    输出: mz_list[50.10775,56.76578,57.0704,69.07013,71.08612,87.50748,103.07556],inten_list[3776,3510,14032,5666,7535,4250,11173]
    '''
    # 将str按照;分割
    strlist = str.split(";")  # [50.10775 3776, 56.76578 3510, ...]
    str_count = len(strlist)
    # 将strlist中的元素按照 分割为两个列表
    mz_list = [] 
    inten_list = []
    for i in range(len(strlist)):
        mz_inten = strlist[i].split(" ")
        try:
            mz, inten = float(mz_inten[0]), float(mz_inten[1])
            mz_list.append(mz)
            inten_list.append(inten)
        except:
            continue
    inten_max = max(inten_list)
    inten_list = [inten/inten_max*1000 for inten in inten_list]
    return mz_list, inten_list

def get_mz_inten_list_type2(str):
    # 将str按照;分割
    strlist = str.split(";")  # [50.10775 3776, 56.76578 3510, ...]
    str_count = len(strlist)
    # 将strlist中的元素按照 分割为两个列表
    mz_list = [] 
    inten_list = []
    for i in range(len(strlist)):
        mz_inten = strlist[i].split(" ")
        try:
            mz, inten = float(mz_inten[0]), float(mz_inten[1])
            mz_list.append(mz)
            inten_list.append(inten)
        except:
            str_count -= 1
            continue
    max_inten = max(inten_list)
    inten_list = [inten/max_inten*1000 for inten in inten_list]
    max_inten = max(inten_list)
    min_inten = min(inten_list)
    if max_inten != min_inten:
        inten_list = [(inten-min_inten)/(max_inten-min_inten) for inten in inten_list]
        count = 0
        for i in range(str_count):
            # 如果峰的强度小于0.05则去除
            if inten_list[i] < 0.05:
                mz_list[i] = 'A'
                inten_list[i] = 'A'
                count += 1
        if count > 5:
            # 去除峰的强度为0的峰
            mz_list = [mz for mz in mz_list if mz != 'A']
            inten_list = [inten for inten in inten_list if inten != 'A']
            inten_list = [inten*(max_inten-min_inten)+min_inten for inten in inten_list]
        else:
            mz_list, inten_list = get_mz_inten_list(str)
    return mz_list, inten_list

def get_ms_diff(mz1, int1, mz2, int2):
    '''
    获取两张谱图的谱图差向量
    输入: mz1, int1, mz2, int2
    输出: ms_diff, 格式为dataframe; 若两张谱图的长度不相等则返回'unknown'
    '''
    if len(mz1) != len(int1) or len(mz2) != len(int2):
        return 'unknown'
    else:
        # 获取ms1丰度排名前5的峰[mz, int]，若少于5个峰则返回全部
        ms1_list = []
        for i in range(len(mz1)):
            ms1_list.append([round(mz1[i],3), round(int1[i],3)])
        ms1_list = sorted(ms1_list, key=lambda x: x[1], reverse=True)
        if len(ms1_list) < 5:
            ms1_list = ms1_list
        else:
            ms1_list = ms1_list[:5]
        # 获取ms2丰度排名前5的峰[mz, int]，若少于5个峰则返回全部
        ms2_list = []
        for i in range(len(mz2)):
            ms2_list.append([round(mz2[i],3), round(int2[i],3)])
        ms2_list = sorted(ms2_list, key=lambda x: x[1], reverse=True)
        if len(ms2_list) < 5:
            ms2_list = ms2_list
        else:
            ms2_list = ms2_list[:5]
        # 计算谱图差向量
        mz_diff_list = []
        ms_diff_list = []
        for a in ms1_list:
            for b in ms2_list:
                mz_diff = abs(a[0] - b[0])
                mz_diff_list.append(round(mz_diff,1)) #四舍五入保留1位小数
                int_avg = (a[1] + b[1]) / 2
                ms_diff_list.append([round(mz_diff,1), int_avg])
        mz_diff_list1 = list(set(mz_diff_list))
        # 计算元素在列表中出现的次数并按出现次数排序，如果出现次数相同则按照丰度排序
        ms_diff_count_list = []
        for i in mz_diff_list1:
            i_int_list = []
            for a in ms_diff_list:
                if i == a[0]:
                    i_int_list.append(a[1])
            i_int_avg = sum(i_int_list) / len(i_int_list)
            ms_diff_count_list.append([i, i_int_avg, mz_diff_list.count(round(i,1))])
        ms_diff_count_list = sorted(ms_diff_count_list, key=lambda x: (x[2], x[1]), reverse=True)
        # 如果超过10个元素则取前10个，如果不足10个元素，则添加0
        ms_diff_final_list = []
        for i in ms_diff_count_list:
            ms_diff_final_list.append(float('%.3f'%i[0]))  #保留3位小数
            ms_diff_final_list.append(float('%.3f'%i[1]))  #保留3位小数
        if len(ms_diff_final_list) < 10:
            for i in range(10 - len(ms_diff_final_list)):
                ms_diff_final_list.append(0)
        else:
            ms_diff_final_list = ms_diff_final_list[:10]
        # list转dataframe
        ms_diff_final_df = pd.DataFrame(np.array(ms_diff_final_list).reshape(1,10).astype(float))
        return ms_diff_final_df

# 判断某个子结构是否在SMILES中，如果在则返回True，否则返回False
def is_substructure(sub_smiles, smiles):
    mol1 = Chem.MolFromSmiles(sub_smiles)
    mol2 = Chem.MolFromSmiles(smiles)
    return mol2.HasSubstructMatch(mol1)

# SMILES计算相对分子质量
def get_mass(smiles):
    mol = Chem.MolFromSmiles(smiles)
    return rdMolDescriptors.CalcExactMolWt(mol)

def get_pairdiff(model_id, mz1, int1, mz2, int2):
    '''
    预测两张谱图的结构差
    输入: model_id, 谱图差向量ms_diff
    输出: 被取代官能团pair_diff1, 取代官能团pair_diff2; 若预测失败则返回'unknown', 'unknown'
    '''
    try:
        ms_diff = get_ms_diff(mz1, int1, mz2, int2)
        if model_id != 'unknown':
            mission_model = joblib.load(r'model\\'+ str(model_id) + r'.joblib')
            label_list = mass_label_dict[model_id]
            y_pred = mission_model.predict(ms_diff)[0]
            pair_diff1, pairdiff2 = label_list[int(y_pred)][0], label_list[int(y_pred)][1]
            # print('pair_diff1:', pair_diff1, 'pair_diff2:', pairdiff2)
        else:
            pair_diff1, pairdiff2 = 'unknown', 'unknown'
        return pair_diff1, pairdiff2
    except:
        print('预测失败')
        return 'unknown', 'unknown'

def get_model_id(mass1, mass2):
    '''
    获取适合的模型id
    输入: 谱图1质量mass1, 谱图2质量mass2
    输出: 适合的模型model_id; 若无适合的模型则返回'unknown'
    '''
    mass_diff = abs(mass1-mass2)
    model_id = 'unknown'
    for model_id0 in model_id_list:
        mass0, mass1 = mass_range_dict[model_id0][0], mass_range_dict[model_id0][1]
        if mass_diff >= mass0 and mass_diff <= mass1:
            model_id = model_id0
            break
    return model_id

def MSMSspe(str1):
    '''
    将str转为两层array
    '''
    count=str1.count(";")
    msmsspe=np.zeros(shape=(count,2))
    strlist=str1.split(";")
    for i in range(count):
        s=strlist[i].split(" ")
        msmsspe[i][0]=eval(s[0])
        msmsspe[i][1]=eval(s[1])
    return msmsspe

def MSMSspe_type2(str1):
    count=str1.count(";")
    msmsspe=np.zeros(shape=(count,2))
    strlist=str1.split(";")
    for i in range(count):
        s=strlist[i].split(" ")
        msmsspe[i][0]=eval(s[0])
        msmsspe[i][1]=eval(s[1])
    # 对第二列元素进行归一化
    int_list = msmsspe[:,1]
    max_int = max(int_list)
    min_int = min(int_list)
    if max_int != min_int:
        for i in range(count):
            msmsspe[i][1] = (msmsspe[i][1]-min_int)/(max_int-min_int)
        # 统计第二列小于0.05元素所在行的个数
        count = 0
        for i in range(len(msmsspe)):
            if msmsspe[i][1] <= 0.05:
                count += 1
        if count > 5:
            msmsspe = msmsspe[msmsspe[:,1] > 0.05]
        # 返回原始值
        for i in range(len(msmsspe)):
            msmsspe[i][1] = msmsspe[i][1]*(max_int-min_int)+min_int
    return msmsspe

def MS_sim(str1,str2,error=0.05):
    '''
    DP算法计算谱图相似性
    '''
    try:
        spec1=MSMSspe(str1)
        spec2=MSMSspe(str2)
        if(len(spec1)==0 or len(spec2)==0):
            return 0
        spec=spec2.copy() 
        alignment=np.zeros(shape=(len(spec1),3))
        alignment[:,0:2]=spec1
        for i in range(len(spec1)):
            match=abs(spec[:,0]-spec1[i,0])
            if(min(match)<=error):
                alignment[i,2]=spec[np.argmin(match),1]
                spec=np.delete(spec,np.argmin(match),axis=0)
            if(len(spec)==0):
                break
        alignment=alignment[alignment[:,2]!=0]
        if(len(alignment)==0):
            return 0
        else:
            return np.dot(alignment[:,1],alignment[:,2])/np.sqrt(((np.dot(spec1[:,1],spec1[:,1]))*(np.dot(spec2[:,1],spec2[:,1]))))
    except:
        return 0

def MS_sim_type2(str1,str2,error=0.01):
    '''
    DP算法计算谱图相似性, 去除杂峰
    '''
    try:
        spec1=MSMSspe_type2(str1)
        spec2=MSMSspe_type2(str2)
        if(len(spec1)==0 or len(spec2)==0):
            return 0
        spec=spec2.copy() 
        alignment=np.zeros(shape=(len(spec1),3))
        alignment[:,0:2]=spec1
        for i in range(len(spec1)):
            match=abs(spec[:,0]-spec1[i,0])
            if(min(match)<=error):
                alignment[i,2]=spec[np.argmin(match),1]
                spec=np.delete(spec,np.argmin(match),axis=0)
            if(len(spec)==0):
                break
        alignment=alignment[alignment[:,2]!=0]
        if(len(alignment)==0):
            return 0
        else:
            return np.dot(alignment[:,1],alignment[:,2])/np.sqrt(((np.dot(spec1[:,1],spec1[:,1]))*(np.dot(spec2[:,1],spec2[:,1]))))
    except:
        return 0

def MS_match_type1(str1,str2,error=0.01):
    '''
    计算谱图的碎片匹配率, 不去杂峰
    '''
    try:
        spec1=MSMSspe(str1)
        spec2=MSMSspe(str2)
        if(len(spec1)==0 or len(spec2)==0):
            return 0
        spec=spec2.copy() 
        alignment=np.zeros(shape=(len(spec1),3))
        alignment[:,0:2]=spec1
        for i in range(len(spec1)):
            match=abs(spec[:,0]-spec1[i,0])
            if(min(match)<=error):
                alignment[i,2]=spec[np.argmin(match),1]
                spec=np.delete(spec,np.argmin(match),axis=0)
            if(len(spec)==0):
                break
        alignment=alignment[alignment[:,2]!=0]
        if(len(alignment)==0):
            return 0
        else:
            return len(alignment)/len(spec1)
    except:
        return 0

def MS_match_type2(str1,str2,error=0.01):
    '''
    计算谱图的碎片匹配率, 不去杂峰
    '''
    try:
        spec1=MSMSspe_type2(str1)
        spec2=MSMSspe_type2(str2)
        if(len(spec1)==0 or len(spec2)==0):
            return 0
        spec=spec2.copy() 
        alignment=np.zeros(shape=(len(spec1),3))
        alignment[:,0:2]=spec1
        for i in range(len(spec1)):
            match=abs(spec[:,0]-spec1[i,0])
            if(min(match)<=error):
                alignment[i,2]=spec[np.argmin(match),1]
                spec=np.delete(spec,np.argmin(match),axis=0)
            if(len(spec)==0):
                break
        alignment=alignment[alignment[:,2]!=0]
        if(len(alignment)==0):
            return 0
        else:
            return len(alignment)/len(spec1)
    except:
        return 0

def get_known_unknown_list(data_dir):
    '''
    获取已知节点列表和未知节点列表
    输入: 数据集路径data_dir, 1:PeakID, 2:SMILES(0为未知节点,1为已知节点), 3:m/z, 4:MS2
    输出: 已知节点列表known_list, 未知节点列表unknown_list
    '''
    wb = load_workbook(data_dir)
    ws = wb.active
    known_dot_list = []
    unknown_dot_list = []
    for i in range(2,ws.max_row):
        DotID0 = ws.cell(i,1).value
        SMILES0 = ws.cell(i,2).value  # 0为未知节点，1为已知节点
        if SMILES0 != 'unknown':
            known_dot_list.append(DotID0)
        else:
            unknown_dot_list.append(DotID0)
    return known_dot_list, unknown_dot_list

def get_pairdiff_list(known_dot_list, unknown_dot_list, data_dir, simlim):
    '''
    根据已知节点和未知节点列表预测结构差
    输入：已知节点列表，未知节点列表，样品信息文件路径，相似性阈值
    输出：[已知节点ID, 未知节点ID, model, sim相似性, pairdiff1-pairdiff2, 边类型]列表, 更新后的已知节点列表, 更新后的未知节点列表
    '''
    wb1 = load_workbook(data_dir)
    ws1 = wb1.active
    known_dot_list1, unknown_dot_list1 = known_dot_list, unknown_dot_list
    pairdiff_list = []
    for i in tqdm(range(2, ws1.max_row+1)):
        DotID1 = ws1.cell(i,1).value
        if DotID1 in known_dot_list:
            mass1, MS1 = ws1.cell(i,3).value, ws1.cell(i,4).value
            mz1_list, int1_list = get_mz_inten_list(MS1)
            for j in range(2, ws1.max_row+1):
                DotID2 = ws1.cell(j,1).value
                if DotID2 in unknown_dot_list:
                    mass2, MS2 = ws1.cell(j,3).value, ws1.cell(j,4).value
                    mz2_list, int2_list = get_mz_inten_list(MS2)
                    model = get_model_id(mass1,mass2)
                    sim = MS_sim(MS1,MS2)
                    if sim >= simlim:
                        if model != 'unknown':
                            pairdiff1, pairdiff2 = get_pairdiff(model, mz1_list, int1_list, mz2_list, int2_list)
                            if pairdiff1 != 'unknown':
                                pairdiff_list.append([DotID1, DotID2, model, sim, pairdiff1+'-'+pairdiff2, 1]) # 1表示结构差边
                            else:
                                pairdiff_list.append([DotID1, DotID2, model, sim, abs(mass2-mass1), 2])        # 2表示质量差边
                        else:
                            pairdiff_list.append([DotID1, DotID2, model, sim, abs(mass2-mass1), 2])
                        known_dot_list1.append(DotID2)    # 更新已知节点列表
                        unknown_dot_list1.remove(DotID2)  # 更新未知节点列表
    # 返回known_dot_list1而不是known_dot_list，是为了保证不改变循环过程中的已知节点列表
    return pairdiff_list, known_dot_list1, unknown_dot_list1

def add_element_to_list(list1, element):
    '''
    给嵌套列表中的每个子列表末尾添加一个元素
    '''
    for i in range(len(list1)):
        list1[i].append(element)
    return list1

def ISRN_spread(data_dir, save_data_dir, count, simlim):
    known_dot_list, unknown_dot_list = get_known_unknown_list(data_dir)
    final_pairdiff_list = []
    a = 0
    while a < count:
        start_len = len(final_pairdiff_list)
        #print('第' + str(a+1) + '次传播...') 
        pairdiff_list, known_dot_list, unknown_dot_list = get_pairdiff_list(known_dot_list, unknown_dot_list, data_dir, simlim)
        pairdiff_list = add_element_to_list(pairdiff_list, a)   # 在每个子列表末尾添加传播轮次
        final_pairdiff_list.extend(pairdiff_list)
        a = a + 1
        #print('已知节点数：' + str(len(known_dot_list)))
        #print('未知节点数：' + str(len(unknown_dot_list)))
        if len(final_pairdiff_list) == start_len:
            break   # 如果迭代后没有出现新的边，则跳出循环
    f = open(save_data_dir,'w',encoding='utf-8',newline='')
    csv_writer = csv.writer(f)
    # 写入表头
    csv_writer.writerow(['DotID1','DotID2','model','sim','pairdiff','edge_type','spread_round'])
    for i in final_pairdiff_list:
        csv_writer.writerow(i) 