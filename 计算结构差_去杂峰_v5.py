from openpyxl import load_workbook
import csv
from tqdm import tqdm
import multiprocessing
from multiprocessing import Process
from AutoML_model_use_function_v5 import *

# 划分并行任务区间
def split_mission(row_id_list, dot_id_MS2_dict, dot_id_mass_dict, data_save_dir, row_id_dot_id_dict, mission_num):
    row_id_list.sort()
    totol_num = len(row_id_list)
    mission_num = min(mission_num, totol_num)
    split_range_num = totol_num // mission_num
    # 将row_id_list划分为mission_num份, 输出划分区间的索引
    split_mission_list = []
    for i in range(mission_num):
        if i == mission_num-1:
            start_index, end_index = i*split_range_num, totol_num
        else:
            start_index, end_index = i*split_range_num, (i+1)*split_range_num
        split_mission_list.append((start_index, end_index, row_id_list, dot_id_MS2_dict, 
                                   dot_id_mass_dict, row_id_dot_id_dict, data_save_dir))
    return split_mission_list

# 并行计算所有分子间的结构差
def get_all_pairdiff(start_end):
    writen_list = []
    start, end, row_id_list, dot_id_MS2_dict, dot_id_mass_dict, row_id_dot_id_dict, data_save_dir = start_end
    row_id_list_0 = row_id_list[start:end]
    for row_id0 in tqdm(row_id_list_0):
        dot_id0 = row_id_dot_id_dict[row_id0][0]
        dot_id1 = row_id_dot_id_dict[row_id0][1]
        sim     = row_id_dot_id_dict[row_id0][2]
        if dot_id0 != dot_id1:
            mass0, mass1 = dot_id_mass_dict[dot_id0], dot_id_mass_dict[dot_id1]
            model_id = get_model_id(mass0, mass1)
            msms0, msms1 = dot_id_MS2_dict[dot_id0], dot_id_MS2_dict[dot_id1]
            mz0, int0 = get_mz_inten_list_type2(msms0)
            mz1, int1 = get_mz_inten_list_type2(msms1)
            pairdiff1, pairdiff2 = get_pairdiff(model_id, mz0, int0, mz1, int1)
            if pairdiff1 != 'unknown' and pairdiff2 != 'unknown':
                pairdiff = pairdiff1 + '-' + pairdiff2
                if mass0 < mass1:
                    writen_list.append([dot_id0, dot_id1, sim, pairdiff])
                else: 
                    writen_list.append([dot_id1, dot_id0, sim, pairdiff])
    f = open(data_save_dir + '\\'+ r'结构差.csv','a',encoding='utf-8',newline='')
    csv_writer = csv.writer(f)
    for i in writen_list:
        csv_writer.writerow(i)

def main(data_save_dir, cpu_num):
    cpu_num = int(cpu_num)
    data_dir0 = data_save_dir + '\\'+ r'msdial_result.xlsx'
    wb1 = load_workbook(data_dir0)
    ws1 = wb1.active
    # 获取dot_id和MS2的字典
    # 获取dot_id和mass的字典
    dot_id_mass_dict = {}
    dot_id_MS2_dict = {}
    for i in range(2, ws1.max_row+1):
        mass, msms = ws1.cell(i,4).value, ws1.cell(i,7).value
        if msms != None:
            dot_id_MS2_dict[ws1.cell(i,1).value] = msms
            dot_id_mass_dict[ws1.cell(i,1).value] = mass
    data_dir1 = data_save_dir + '\\'+ r'相似性.xlsx'
    wb = load_workbook(data_dir1)
    ws = wb.active
    # 获取row_id和[dot_id, dot_id, sim]的字典
    row_id_dot_id_dict = {}
    for i in range(1, ws.max_row+1):
        row_id_dot_id_dict[i] = [ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,3).value]
    # 获取键列表
    row_id_list = list(row_id_dot_id_dict.keys())
    
    split_mission_list = split_mission(row_id_list, dot_id_MS2_dict, dot_id_mass_dict, data_save_dir, 
                                       row_id_dot_id_dict, cpu_num)
    pool = multiprocessing.Pool(cpu_num)
    rel = pool.map(get_all_pairdiff, split_mission_list)
    pool.close()

if __name__ == '__main__':
    main()
    