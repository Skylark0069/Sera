from openpyxl import load_workbook
import csv
from tqdm import tqdm
import multiprocessing
from multiprocessing import Process
from AutoML_model_use_function_v5 import *


# 划分并行任务区间
def split_mission(dot_id_list, dot_id_MS2_dict, dot_id_mass_dict, data_save_dir, mission_num, sim_threshold):
    dot_id_list.sort()
    totol_num = len(dot_id_list)
    mission_num = int(mission_num)
    split_range_num = totol_num // mission_num
    # 将dot_id_list划分为mission_num份, 输出划分区间的索引
    split_mission_list = []
    for i in range(mission_num):
        if i == mission_num-1:
            start_index, end_index = i*split_range_num, totol_num
        else:
            start_index, end_index = i*split_range_num, (i+1)*split_range_num  
        split_mission_list.append((start_index, end_index, dot_id_list, dot_id_MS2_dict, dot_id_mass_dict, data_save_dir, sim_threshold))
    return split_mission_list

# 并行计算所有分子间的相似性
def get_all_sim(start_end):
    writen_list = []
    start, end, dot_id_list, dot_id_MS2_dict, dot_id_mass_dict, data_save_dir, sim_threshold = start_end
    dot_id_list_0 = dot_id_list[start:end]
    for dot_id0 in tqdm(dot_id_list_0):
        msms0 = dot_id_MS2_dict[dot_id0]
        mass0 = dot_id_mass_dict[dot_id0]
        for dot_id1 in dot_id_list:
            if dot_id0 != dot_id1:
                msms1 = dot_id_MS2_dict[dot_id1]
                mass1 = dot_id_mass_dict[dot_id1]
                if abs(mass0-mass1) < 100:  # 100是结构差模型的最大质量差
                    sim = MS_sim(msms0, msms1)
                    if sim > sim_threshold:
                        if mass0 < mass1:
                            writen_list.append([dot_id0, dot_id1, sim])
                        else: 
                            writen_list.append([dot_id1, dot_id0, sim])
    f = open(data_save_dir + '\\'+ r'相似性.csv','a',encoding='utf-8',newline='')
    csv_writer = csv.writer(f)
    for i in writen_list:
        csv_writer.writerow(i)

def main(data_save_dir, cpu_num, sim_threshold):
    cpu_num = int(cpu_num)
    wb = load_workbook(data_save_dir + '\\'+ r'msdial_result.xlsx')
    ws = wb.active
    # 获取dot_id和MS2的字典
    # 获取dot_id和mass的字典
    dot_id_mass_dict = {}
    dot_id_MS2_dict = {}
    for i in range(2, ws.max_row+1):
        msms_str = ws.cell(i, 7).value
        if msms_str != None:
            dot_id_MS2_dict[ws.cell(i,1).value] = ws.cell(i,7).value
            dot_id_mass_dict[ws.cell(i,1).value] = float(ws.cell(i,4).value)
    # 获取键列表
    dot_id_list = list(dot_id_MS2_dict.keys())
    
    split_mission_list = split_mission(dot_id_list, dot_id_MS2_dict, dot_id_mass_dict, data_save_dir, cpu_num, sim_threshold)
    pool = multiprocessing.Pool(cpu_num)
    rel = pool.map(get_all_sim, split_mission_list)
    pool.close()

if __name__ == '__main__':
    main()