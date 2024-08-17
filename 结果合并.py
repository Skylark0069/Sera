from tqdm import tqdm
from openpyxl import load_workbook
import csv

def main1(data_save_dir, sim_threshold):
    wb1 = load_workbook(data_save_dir + '\\'+ r'msdial_result.xlsx')
    ws1 = wb1.active
    # 建立dot_id和SMILES的字典
    dot_id_SMILES_dict = {}
    dot_id_mass_dict = {}
    for i in tqdm(range(2, ws1.max_row+1)):
        dot_id = int(ws1.cell(i, 1).value)
        SMILES = ws1.cell(i, 5).value
        mass = float(ws1.cell(i, 4).value)
        dot_id_mass_dict[dot_id] = mass
        if SMILES != "unknown":
            dot_id_SMILES_dict[dot_id] = SMILES
    
    final_list = []
    # 读取结构差结果
    wb2 = load_workbook(data_save_dir + '\\'+ r'结构差.xlsx')
    ws2 = wb2.active
    for i in tqdm(range(1, ws2.max_row+1)):
        dot_id1, dot_id2, sim, pairdiff = int(ws2.cell(i,1).value), int(ws2.cell(i,2).value), float(ws2.cell(i,3).value), ws2.cell(i,4).value
        try:
            SMILES1 = dot_id_SMILES_dict[dot_id1]
        except:
            SMILES1 = "unknown"
        try:
            SMILES2 = dot_id_SMILES_dict[dot_id2]
        except:
            SMILES2 = "unknown"
        if sim > sim_threshold:
            if SMILES1 == "unknown" or SMILES2 == "unknown":
                final_list.append([dot_id1, SMILES1, dot_id2, SMILES2, sim, pairdiff])

    # 读取KU网络的结果
    wb3 = load_workbook(data_save_dir + '\\'+ r'KU_net.xlsx')
    ws3 = wb3.active
    for i in tqdm(range(1, ws3.max_row+1)):
        dot_id1, dot_id2 = ws3.cell(i,1).value, ws3.cell(i,3).value
        if dot_id1 != "nist":
            dot_id1, smiles1 = int(ws3.cell(i,1).value), ws3.cell(i,2).value
        else: 
            dot_id1, smiles1 = ws3.cell(i,1).value, ws3.cell(i,2).value
        if dot_id2 != "nist":
            dot_id2, smiles2 = int(ws3.cell(i,3).value), ws3.cell(i,4).value
        else:
            dot_id2, smiles2 = ws3.cell(i,3).value, ws3.cell(i,4).value
        sim, pairdiff = float(ws3.cell(i,5).value), ws3.cell(i,6).value
        if sim > sim_threshold:
            final_list.append([dot_id1, smiles1, dot_id2, smiles2, sim, pairdiff])
    
    # 读取真实结构差结果
    try:
        wb4 = load_workbook(data_save_dir + '\\'+ r'真实结构差.xlsx')
        ws4 = wb4.active
        for i in tqdm(range(1, ws4.max_row+1)):
            dot_id1, dot_id2 = int(ws4.cell(i,1).value), int(ws4.cell(i,2).value)
            smiles1, smiles2 = dot_id_SMILES_dict[dot_id1], dot_id_SMILES_dict[dot_id2]
            pairdiff = ws4.cell(i,3).value
            final_list.append([dot_id1, smiles1, dot_id2, smiles2, 1, pairdiff])
    except:
        pass

    final_list_x = []
    # 标记其中的同分异构体
    for x_list in tqdm(final_list):
        dot_id1, smiles1, dot_id2, smiles2, sim, pairdiff = x_list
        if dot_id1 != "nist" and dot_id2 != "nist":
            mass1, mass2 = dot_id_mass_dict[dot_id1], dot_id_mass_dict[dot_id2]
            if abs(mass1 - mass2) <= 0.05:
                final_list_x.append([dot_id1, smiles1, dot_id2, smiles2, sim, "Isomers"])
            else:
                final_list_x.append([dot_id1, smiles1, dot_id2, smiles2, sim, pairdiff])
        else:
            final_list_x.append([dot_id1, smiles1, dot_id2, smiles2, sim, pairdiff])
            
    # 写入csv文件
    with open(data_save_dir + '\\'+ r'总表.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(final_list_x)

if __name__ == "__main__":
    main1()